import os, re, json
import gradio as gr
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook

USERS_XLSX = "usersinfo.xlsx"
USER_DATA_DIR = "user_data"

# ===== Helpers =====
def ensure_users_xlsx():
    if not os.path.exists(USERS_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "users"
        ws.append(["username", "password_hash", "favorite"])
        wb.save(USERS_XLSX)

def get_user(username):
    if not os.path.exists(USERS_XLSX): return None
    wb = load_workbook(USERS_XLSX)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return {"username": row[0], "password_hash": row[1], "favorite": row[2]}
    return None

def add_user(username, password_hash, favorite):
    wb = load_workbook(USERS_XLSX)
    ws = wb.active
    ws.append([username, password_hash, favorite])
    wb.save(USERS_XLSX)

def user_folder(username):
    return os.path.join(USER_DATA_DIR, username)

def ensure_user_folder(username):
    os.makedirs(user_folder(username), exist_ok=True)

def safe_filename(s):
    return re.sub(r"[^A-Za-z0-9_-]", "_", s.strip()) or "note"

# ===== Auth + Notes Logic =====
current_user = None

def register(username, password, favorite):
    global current_user
    if get_user(username):
        return "❌ Username already exists", None
    pw_hash = generate_password_hash(password)
    add_user(username, pw_hash, favorite)
    ensure_user_folder(username)
    current_user = username
    return f"✅ Registered & logged in as {username}", username

def login(username, password):
    global current_user
    user = get_user(username)
    if not user or not check_password_hash(user["password_hash"], password):
        return "❌ Invalid login", None
    current_user = username
    return f"✅ Logged in as {username}", username

def logout():
    global current_user
    current_user = None
    return "👋 Logged out", None

def create_note(title, content, ntype="text"):
    if not current_user:
        return "❌ Not logged in"
    safe = safe_filename(title)
    folder = user_folder(current_user)
    ensure_user_folder(current_user)
    filepath = os.path.join(folder, f"{safe}_{ntype}.json")
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump({"title": title, "content": content, "type": ntype}, f)
    return f"✅ Note '{title}' saved!"

def list_notes():
    if not current_user: return []
    folder = user_folder(current_user)
    notes = []
    if os.path.exists(folder):
        for f in os.listdir(folder):
            if f.endswith(".json"):
                with open(os.path.join(folder, f), "r", encoding="utf-8") as fh:
                    notes.append(json.load(fh))
    return notes

def delete_note(title):
    if not current_user: return "❌ Not logged in"
    folder = user_folder(current_user)
    safe = safe_filename(title)
    for f in os.listdir(folder):
        if f.startswith(safe):
            os.remove(os.path.join(folder, f))
            return f"🗑 Deleted '{title}'"
    return "❌ Note not found"

# ===== Gradio UI =====
with gr.Blocks(title="📝 My Notes App") as demo:
    gr.Markdown("# 📝 My Notes App\nA simple Keep Notes clone with login + notes!")

    with gr.Tab("Register / Login"):
        r_user = gr.Textbox(label="Register Username")
        r_pass = gr.Textbox(label="Register Password", type="password")
        r_fav = gr.Textbox(label="Favorite Game/Dish")
        reg_btn = gr.Button("Register")
        reg_status = gr.Textbox(label="Status", interactive=False)

        l_user = gr.Textbox(label="Login Username")
        l_pass = gr.Textbox(label="Login Password", type="password")
        login_btn = gr.Button("Login")
        login_status = gr.Textbox(label="Status", interactive=False)

        logout_btn = gr.Button("Logout")
        logout_status = gr.Textbox(label="Status", interactive=False)

        reg_btn.click(register, [r_user, r_pass, r_fav], [reg_status, login_status])
        login_btn.click(login, [l_user, l_pass], [login_status, login_status])
        logout_btn.click(logout, [], [logout_status, login_status])

    with gr.Tab("Notes"):
        note_title = gr.Textbox(label="Title")
        note_content = gr.Textbox(label="Content", lines=5)
        save_btn = gr.Button("Save Note")
        save_status = gr.Textbox(label="Status", interactive=False)

        notes_list = gr.Dataframe(headers=["Title", "Type", "Content"], interactive=False)
        refresh_btn = gr.Button("🔄 Refresh Notes")
        del_title = gr.Textbox(label="Delete Note by Title")
        del_btn = gr.Button("Delete Note")
        del_status = gr.Textbox(label="Status", interactive=False)

        save_btn.click(create_note, [note_title, note_content], save_status)
        refresh_btn.click(lambda: [(n["title"], n["type"], n["content"]) for n in list_notes()], [], notes_list)
        del_btn.click(delete_note, [del_title], del_status)

# ===== Launch =====
if __name__ == "__main__":
    ensure_users_xlsx()
    os.makedirs(USER_DATA_DIR, exist_ok=True)
    demo.launch(share=True, server_name="0.0.0.0", server_port=7860)
